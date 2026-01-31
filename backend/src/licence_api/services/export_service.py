"""Export service for CSV and Excel generation."""

import csv
import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any
from uuid import UUID

from sqlalchemy.ext.asyncio import AsyncSession

from licence_api.repositories.license_repository import LicenseRepository
from licence_api.repositories.cost_snapshot_repository import CostSnapshotRepository


class ExportService:
    """Service for exporting data to CSV and Excel formats."""

    def __init__(self, session: AsyncSession) -> None:
        """Initialize service with database session."""
        self.session = session
        self.license_repo = LicenseRepository(session)
        self.snapshot_repo = CostSnapshotRepository(session)

    async def export_licenses_csv(
        self,
        provider_id: UUID | None = None,
        department: str | None = None,
        status: str | None = None,
    ) -> str:
        """Export licenses to CSV format.

        Args:
            provider_id: Filter by provider
            department: Filter by department
            status: Filter by status

        Returns:
            CSV string
        """
        results, _ = await self.license_repo.get_all_with_details(
            provider_id=provider_id,
            department=department,
            status=status,
            limit=50000,
        )

        output = io.StringIO()
        writer = csv.writer(output)

        # Header row
        writer.writerow([
            "License ID",
            "Provider",
            "External User ID",
            "License Type",
            "Status",
            "Employee Name",
            "Employee Email",
            "Employee Department",
            "Monthly Cost",
            "Currency",
            "Last Activity",
            "Assigned At",
            "Synced At",
        ])

        # Data rows
        for lic, provider, employee in results:
            writer.writerow([
                str(lic.id),
                provider.display_name,
                lic.external_user_id,
                lic.license_type or "",
                lic.status,
                employee.full_name if employee else "",
                employee.email if employee else "",
                employee.department if employee else "",
                str(lic.monthly_cost) if lic.monthly_cost else "",
                lic.currency or "EUR",
                lic.last_activity_at.isoformat() if lic.last_activity_at else "",
                lic.assigned_at.isoformat() if lic.assigned_at else "",
                lic.synced_at.isoformat() if lic.synced_at else "",
            ])

        return output.getvalue()

    async def export_costs_csv(
        self,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> str:
        """Export cost data to CSV format.

        Args:
            start_date: Start date for the export
            end_date: End date for the export

        Returns:
            CSV string
        """
        if start_date is None:
            start_date = date.today().replace(day=1, month=1)
        if end_date is None:
            end_date = date.today()

        snapshots = await self.snapshot_repo.get_range(
            start_date=start_date,
            end_date=end_date,
            provider_id=None,
        )

        output = io.StringIO()
        writer = csv.writer(output)

        # Header row
        writer.writerow([
            "Date",
            "Total Cost",
            "License Count",
            "Active Count",
            "Unassigned Count",
            "Currency",
        ])

        # Data rows
        for snapshot in snapshots:
            writer.writerow([
                snapshot.snapshot_date.isoformat(),
                str(snapshot.total_cost),
                snapshot.license_count,
                snapshot.active_count,
                snapshot.unassigned_count,
                snapshot.currency,
            ])

        # If no snapshots, add current data
        if not snapshots:
            results, _ = await self.license_repo.get_all_with_details(limit=50000)

            total_cost = Decimal("0")
            total_count = 0
            active_count = 0
            unassigned_count = 0

            for lic, provider, employee in results:
                if provider.name == "hibob":
                    continue
                total_count += 1
                if lic.monthly_cost:
                    total_cost += lic.monthly_cost
                if lic.status == "active":
                    active_count += 1
                if employee is None:
                    unassigned_count += 1

            writer.writerow([
                date.today().isoformat(),
                str(total_cost),
                total_count,
                active_count,
                unassigned_count,
                "EUR",
            ])

        return output.getvalue()

    async def export_full_report_excel(self) -> bytes:
        """Export full report to Excel format with multiple sheets.

        Returns:
            Excel file bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")

        wb = Workbook()

        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Get license statistics
        results, total = await self.license_repo.get_all_with_details(limit=50000)

        total_cost = Decimal("0")
        active_count = 0
        unassigned_count = 0
        provider_stats: dict[str, dict[str, Any]] = {}

        for lic, provider, employee in results:
            if provider.name == "hibob":
                continue

            if lic.monthly_cost:
                total_cost += lic.monthly_cost
            if lic.status == "active":
                active_count += 1
            if employee is None:
                unassigned_count += 1

            # Aggregate by provider
            pname = provider.display_name
            if pname not in provider_stats:
                provider_stats[pname] = {"count": 0, "cost": Decimal("0")}
            provider_stats[pname]["count"] += 1
            if lic.monthly_cost:
                provider_stats[pname]["cost"] += lic.monthly_cost

        # Style definitions
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # Summary content
        ws_summary["A1"] = "License Report Summary"
        ws_summary["A1"].font = Font(bold=True, size=14)

        ws_summary["A3"] = "Generated"
        ws_summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")

        ws_summary["A5"] = "Total Licenses"
        ws_summary["B5"] = total

        ws_summary["A6"] = "Active Licenses"
        ws_summary["B6"] = active_count

        ws_summary["A7"] = "Unassigned Licenses"
        ws_summary["B7"] = unassigned_count

        ws_summary["A8"] = "Total Monthly Cost"
        ws_summary["B8"] = f"EUR {total_cost:,.2f}"

        # Provider breakdown
        ws_summary["A10"] = "Provider"
        ws_summary["B10"] = "Licenses"
        ws_summary["C10"] = "Monthly Cost"
        ws_summary["A10"].font = header_font
        ws_summary["B10"].font = header_font
        ws_summary["C10"].font = header_font
        ws_summary["A10"].fill = header_fill
        ws_summary["B10"].fill = header_fill
        ws_summary["C10"].fill = header_fill

        row = 11
        for pname, stats in sorted(provider_stats.items()):
            ws_summary[f"A{row}"] = pname
            ws_summary[f"B{row}"] = stats["count"]
            ws_summary[f"C{row}"] = f"EUR {stats['cost']:,.2f}"
            row += 1

        # Adjust column widths
        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 15
        ws_summary.column_dimensions["C"].width = 18

        # Licenses Sheet
        ws_licenses = wb.create_sheet("Licenses")

        # Headers
        license_headers = [
            "License ID", "Provider", "External User ID", "License Type", "Status",
            "Employee Name", "Employee Email", "Department", "Monthly Cost", "Currency",
            "Last Activity", "Assigned At"
        ]

        for col, header in enumerate(license_headers, 1):
            cell = ws_licenses.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Data
        row = 2
        for lic, provider, employee in results:
            if provider.name == "hibob":
                continue

            ws_licenses.cell(row=row, column=1, value=str(lic.id))
            ws_licenses.cell(row=row, column=2, value=provider.display_name)
            ws_licenses.cell(row=row, column=3, value=lic.external_user_id)
            ws_licenses.cell(row=row, column=4, value=lic.license_type or "")
            ws_licenses.cell(row=row, column=5, value=lic.status)
            ws_licenses.cell(row=row, column=6, value=employee.full_name if employee else "")
            ws_licenses.cell(row=row, column=7, value=employee.email if employee else "")
            ws_licenses.cell(row=row, column=8, value=employee.department if employee else "")
            ws_licenses.cell(row=row, column=9, value=float(lic.monthly_cost) if lic.monthly_cost else None)
            ws_licenses.cell(row=row, column=10, value=lic.currency or "EUR")
            ws_licenses.cell(row=row, column=11, value=lic.last_activity_at.strftime("%Y-%m-%d") if lic.last_activity_at else "")
            ws_licenses.cell(row=row, column=12, value=lic.assigned_at.strftime("%Y-%m-%d") if lic.assigned_at else "")
            row += 1

        # Adjust column widths
        for col in range(1, len(license_headers) + 1):
            ws_licenses.column_dimensions[chr(64 + col)].width = 18

        # Costs Sheet
        ws_costs = wb.create_sheet("Costs")

        # Get cost snapshots
        today = date.today()
        start = today.replace(day=1, month=1)
        snapshots = await self.snapshot_repo.get_range(start_date=start, end_date=today)

        cost_headers = ["Date", "Total Cost", "License Count", "Active Count", "Unassigned Count"]

        for col, header in enumerate(cost_headers, 1):
            cell = ws_costs.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        if snapshots:
            row = 2
            for snapshot in snapshots:
                ws_costs.cell(row=row, column=1, value=snapshot.snapshot_date.strftime("%Y-%m-%d"))
                ws_costs.cell(row=row, column=2, value=float(snapshot.total_cost))
                ws_costs.cell(row=row, column=3, value=snapshot.license_count)
                ws_costs.cell(row=row, column=4, value=snapshot.active_count)
                ws_costs.cell(row=row, column=5, value=snapshot.unassigned_count)
                row += 1
        else:
            # Current data if no snapshots
            ws_costs.cell(row=2, column=1, value=today.strftime("%Y-%m-%d"))
            ws_costs.cell(row=2, column=2, value=float(total_cost))
            ws_costs.cell(row=2, column=3, value=total)
            ws_costs.cell(row=2, column=4, value=active_count)
            ws_costs.cell(row=2, column=5, value=unassigned_count)

        # Adjust column widths
        for col in range(1, len(cost_headers) + 1):
            ws_costs.column_dimensions[chr(64 + col)].width = 18

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
